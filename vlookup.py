'''
刚开始学python，没有复用，也很不健壮，但是能基本实现提取数据
就这样把，不修改了
等到学会做界面交互，再来把这个功能打包，更希望能是做在网站上
借来下做个能修改word格式的
希望这是一个好的开始
万事大吉
'''

#import pandas as pd
import openpyxl
from openpyxl import Workbook
#filename1 = input('')
#filename2 = input('')

#正式的版本，把文件名改为变量
#result1 = pd.read_excel('x.xlsx')
#result2 = pd.read_excel('ex.xlsx')
wb = openpyxl.load_workbook('x.xlsx')
sheet = wb['Sheet1']
wb2 = openpyxl.load_workbook('ex.xlsx')
sheet2 = wb2['Sheet1']
newsheet = {}
newsheet2 = {}

for row in range(2,sheet.max_row+1):
    name = sheet['A'+str(row)].value
    xinxi = sheet['C'+str(row)].value

    newsheet.setdefault(name,xinxi)

for row2 in range(2,sheet2.max_row+1):
    name2 = sheet2['A'+str(row2)].value
    xinxi2 = sheet2['C'+str(row2)].value

    newsheet2.setdefault(name2,xinxi2)

#print(newsheet.keys())
#for i in newsheet.keys():
    #print(i)
#print(list(newsheet.keys()))
#print(list(newsheet2.keys()))
liebiao = list(newsheet.keys())
liebiao2 = list(newsheet2.keys())
spam = []
for i in range(0,len(liebiao)):
    for y in range(0,len(liebiao2)):
        if liebiao[i] == liebiao2[y]:
            spam.append(liebiao[i])

#for k in range(0,len(spam)):
    #print(spam[k] )
    #print(newsheet[spam[k]])

f = Workbook()
table = f.active
table.title = 'Sheet1'

for k in range(1,len(spam)):
    table.cell(row=k+1,column=1).value = spam[k]
    table.cell(row=k+1,column=2).value = newsheet[spam[k]]

#print(table.cell(row=1,column=1).value)
f.save('1.xlsx')
f.close()



#print(len(result1['地区']))
#len1 = len(result1['姓名'])
#len2 = len(result2['姓名'])
#for i in range(0,len1):
    #print(result1.iloc[i,1])
    #data1 = result1.iloc[i,1]
    #print(data1)

    #for y in range(0,len2):
        #data2 = result2.iloc[y, 0]
        #print(data2)

        #if data1 == data2:
            #newsheet.setdefault()
            #print(newsheet)




